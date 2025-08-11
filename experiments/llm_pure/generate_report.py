#!/usr/bin/env python3
"""
Automatic Excel Report Generator for LLM Pure Experiments

Generates comprehensive comparison reports after experiments are complete.
"""

import json
import pandas as pd
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Tuple

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "src"))
from llm_pure.config import config


class ExperimentReportGenerator:
    """Generate Excel reports comparing different experiment results"""
    
    def __init__(self):
        self.results_dir = config.results_dir
        self.reports_dir = self.results_dir / "reports"
        self.reports_dir.mkdir(exist_ok=True)
        self.security_smells = [
            "Admin by default",
            "Empty password", 
            "Hard-coded secret",
            "Missing default in case statement",
            "No integrity check",
            "Suspicious comment",
            "Unrestricted IP address",
            "Use of HTTP without SSL/TLS",
            "Use of weak cryptography algorithms"
        ]
    
    def find_latest_experiments(self, iac_tech: str = "puppet", limit: int = 5) -> List[str]:
        """Find the latest experiment IDs for the specified technology"""
        experiment_files = []
        
        # Find batch files for the technology
        for file_path in self.results_dir.glob(f"batch_{iac_tech}_*.json"):
            experiment_id = file_path.stem.replace(f"batch_{iac_tech}_", "")
            experiment_files.append((experiment_id, file_path))
        
        # Sort by experiment ID (timestamp) and return latest
        experiment_files.sort(key=lambda x: x[0], reverse=True)
        return [exp_id for exp_id, _ in experiment_files[:limit]]
    
    def load_experiment_results(self, experiment_id: str, iac_tech: str = "puppet") -> Dict[str, Any]:
        """Load detailed results for a specific experiment"""
        batch_file = self.results_dir / f"batch_{iac_tech}_{experiment_id}.json"
        
        if not batch_file.exists():
            raise FileNotFoundError(f"Batch file not found: {batch_file}")
        
        with open(batch_file, 'r') as f:
            data = json.load(f)
        
        return {
            'experiment_id': experiment_id,
            'model_name': data.get('model_name', 'Unknown'),
            'prompt_style': data.get('processing_metadata', {}).get('prompt_style', 'Unknown'),
            'iac_technology': data.get('iac_technology', iac_tech),
            'total_files': data.get('total_files', 0),
            'successful_files': data.get('successful_files', 0),
            'overall_metrics': data.get('overall_metrics', {}),
            'error_analysis': data.get('overall_error_analysis', {}),
            'timestamp': data.get('processing_metadata', {}).get('timestamp', 'Unknown')
        }
    
    def extract_smell_metrics(self, overall_metrics: Dict) -> Dict[str, Dict]:
        """Extract per-smell metrics from overall metrics"""
        by_category = overall_metrics.get('by_category', {})
        
        smell_metrics = {}
        for smell in self.security_smells:
            if smell in by_category:
                metrics = by_category[smell]
                smell_metrics[smell] = {
                    'precision': metrics.get('precision', 0.0),
                    'recall': metrics.get('recall', 0.0),
                    'f1': metrics.get('f1', 0.0),
                    'true_positives': metrics.get('true_positives', 0),
                    'false_positives': metrics.get('false_positives', 0),
                    'false_negatives': metrics.get('false_negatives', 0),
                    'support': metrics.get('support', 0)
                }
            else:
                # No data for this smell
                smell_metrics[smell] = {
                    'precision': 0.0,
                    'recall': 0.0,
                    'f1': 0.0,
                    'true_positives': 0,
                    'false_positives': 0,
                    'false_negatives': 0,
                    'support': 0
                }
        
        return smell_metrics
    
    def create_comparison_table(self, experiments: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create comparison table in GLITCH paper format"""
        
        # Initialize the comparison data
        comparison_data = {
            'Smell Name': self.security_smells + ['Average']
        }
        
        # Add columns for each experiment
        for exp in experiments:
            prompt_style = exp['prompt_style']
            smell_metrics = self.extract_smell_metrics(exp['overall_metrics'])
            
            # Precision column
            precisions = []
            recalls = []
            f1s = []
            
            for smell in self.security_smells:
                metrics = smell_metrics.get(smell, {})
                precisions.append(metrics.get('precision', 0.0))
                recalls.append(metrics.get('recall', 0.0))
                f1s.append(metrics.get('f1', 0.0))
            
            # Calculate averages (non-zero values only for more meaningful averages)
            avg_precision = sum(p for p in precisions if p > 0) / len([p for p in precisions if p > 0]) if any(p > 0 for p in precisions) else 0.0
            avg_recall = sum(r for r in recalls if r > 0) / len([r for r in recalls if r > 0]) if any(r > 0 for r in recalls) else 0.0
            avg_f1 = sum(f for f in f1s if f > 0) / len([f for f in f1s if f > 0]) if any(f > 0 for f in f1s) else 0.0
            
            precisions.append(avg_precision)
            recalls.append(avg_recall)
            f1s.append(avg_f1)
            
            comparison_data[f'{prompt_style}_Precision'] = precisions
            comparison_data[f'{prompt_style}_Recall'] = recalls
            comparison_data[f'{prompt_style}_F1'] = f1s
        
        return pd.DataFrame(comparison_data)
    
    def create_detailed_metrics_table(self, experiments: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create detailed metrics table with additional statistics"""
        
        detailed_data = []
        
        for exp in experiments:
            overall = exp['overall_metrics'].get('overall', {})
            smell_metrics = self.extract_smell_metrics(exp['overall_metrics'])
            
            # Overall experiment info
            base_info = {
                'Experiment_ID': exp['experiment_id'],
                'Prompt_Style': exp['prompt_style'],
                'Model': exp['model_name'],
                'Total_Files': exp['total_files'],
                'Successful_Files': exp['successful_files'],
                'Overall_Precision': overall.get('precision', 0.0),
                'Overall_Recall': overall.get('recall', 0.0),
                'Overall_F1': overall.get('f1', 0.0),
                'Total_TP': overall.get('true_positives', 0),
                'Total_FP': overall.get('false_positives', 0),
                'Total_FN': overall.get('false_negatives', 0),
                'Timestamp': exp['timestamp']
            }
            
            # Add per-smell metrics
            for smell in self.security_smells:
                metrics = smell_metrics.get(smell, {})
                smell_key = smell.replace(' ', '_').replace('-', '_').replace('/', '_')
                base_info.update({
                    f'{smell_key}_Precision': metrics.get('precision', 0.0),
                    f'{smell_key}_Recall': metrics.get('recall', 0.0),
                    f'{smell_key}_F1': metrics.get('f1', 0.0),
                    f'{smell_key}_Support': metrics.get('support', 0)
                })
            
            detailed_data.append(base_info)
        
        return pd.DataFrame(detailed_data)
    
    def create_glitch_comparison_table(self, experiments: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create table in exact GLITCH paper format for easy comparison"""
        
        # GLITCH baseline data from the paper (for reference)
        glitch_data = {
            'Admin by default': {'precision': 0.81, 'recall': 0.93},
            'Empty password': {'precision': 1.0, 'recall': 1.0},
            'Hard-coded secret': {'precision': 0.14, 'recall': 0.82},
            'Unrestricted IP address': {'precision': 1.0, 'recall': 1.0},  # Invalid IP address binding
            'Suspicious comment': {'precision': 0.39, 'recall': 1.0},
            'Use of HTTP without SSL/TLS': {'precision': 0.45, 'recall': 1.0},
            'No integrity check': {'precision': 0.0, 'recall': 0.0},  # N/D in paper
            'Use of weak cryptography algorithms': {'precision': 0.57, 'recall': 1.0},
            'Missing default in case statement': {'precision': 0.83, 'recall': 1.0}
        }
        
        # Create table structure
        table_data = []
        
        for smell in self.security_smells:
            row = {'Smell Name': smell}
            
            # Add GLITCH baseline (for reference)
            glitch_metrics = glitch_data.get(smell, {'precision': 0.0, 'recall': 0.0})
            row['GLITCH_Precision'] = glitch_metrics['precision']
            row['GLITCH_Recall'] = glitch_metrics['recall']
            
            # Add our experiments
            for exp in experiments:
                prompt_style = exp['prompt_style']
                smell_metrics = self.extract_smell_metrics(exp['overall_metrics'])
                metrics = smell_metrics.get(smell, {})
                
                row[f'{prompt_style}_Precision'] = metrics.get('precision', 0.0)
                row[f'{prompt_style}_Recall'] = metrics.get('recall', 0.0)
            
            table_data.append(row)
        
        # Add average row
        avg_row = {'Smell Name': 'Average'}
        
        # GLITCH average
        glitch_precisions = [glitch_data[smell]['precision'] for smell in self.security_smells if smell in glitch_data]
        glitch_recalls = [glitch_data[smell]['recall'] for smell in self.security_smells if smell in glitch_data]
        avg_row['GLITCH_Precision'] = sum(glitch_precisions) / len(glitch_precisions)
        avg_row['GLITCH_Recall'] = sum(glitch_recalls) / len(glitch_recalls)
        
        # Our experiments' averages
        for exp in experiments:
            prompt_style = exp['prompt_style']
            overall = exp['overall_metrics'].get('overall', {})
            
            avg_row[f'{prompt_style}_Precision'] = overall.get('precision', 0.0)
            avg_row[f'{prompt_style}_Recall'] = overall.get('recall', 0.0)
        
        table_data.append(avg_row)
        
        return pd.DataFrame(table_data)
    
    def generate_excel_report(self, experiment_ids: List[str], iac_tech: str = "puppet", output_file: str = None) -> str:
        """Generate comprehensive Excel report"""
        
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = self.reports_dir / f"comparison_report_{iac_tech}_{timestamp}.xlsx"
        
        # Load experiment data
        experiments = []
        for exp_id in experiment_ids:
            try:
                exp_data = self.load_experiment_results(exp_id, iac_tech)
                experiments.append(exp_data)
                print(f"✓ Loaded experiment {exp_id} ({exp_data['prompt_style']})")
            except FileNotFoundError as e:
                print(f"❌ Failed to load experiment {exp_id}: {e}")
                continue
        
        if not experiments:
            raise ValueError("No valid experiments found")
        
        # Create different views of the data
        comparison_df = self.create_comparison_table(experiments)
        detailed_df = self.create_detailed_metrics_table(experiments)
        glitch_comparison_df = self.create_glitch_comparison_table(experiments)
        
        # Create experiment summary
        summary_data = []
        for exp in experiments:
            overall = exp['overall_metrics'].get('overall', {})
            summary_data.append({
                'Experiment_ID': exp['experiment_id'],
                'Prompt_Style': exp['prompt_style'],
                'Model': exp['model_name'],
                'Technology': exp['iac_technology'],
                'Files_Processed': f"{exp['successful_files']}/{exp['total_files']}",
                'Overall_Precision': f"{overall.get('precision', 0.0):.3f}",
                'Overall_Recall': f"{overall.get('recall', 0.0):.3f}",
                'Overall_F1': f"{overall.get('f1', 0.0):.3f}",
                'True_Positives': overall.get('true_positives', 0),
                'False_Positives': overall.get('false_positives', 0),
                'False_Negatives': overall.get('false_negatives', 0),
                'Timestamp': exp['timestamp']
            })
        
        summary_df = pd.DataFrame(summary_data)
        
        # Write to Excel with multiple sheets
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Summary sheet
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # GLITCH comparison (main result)
            glitch_comparison_df.to_excel(writer, sheet_name='GLITCH_Comparison', index=False)
            
            # Detailed comparison
            comparison_df.to_excel(writer, sheet_name='Detailed_Comparison', index=False)
            
            # Full metrics
            detailed_df.to_excel(writer, sheet_name='Full_Metrics', index=False)
            
            # Format the sheets
            self._format_excel_sheets(writer, glitch_comparison_df, comparison_df)
        
        print(f"📊 Excel report generated: {output_file}")
        return str(output_file)
    
    def _format_excel_sheets(self, writer, glitch_df, comparison_df):
        """Apply formatting to Excel sheets"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Format GLITCH comparison sheet
        ws = writer.sheets['GLITCH_Comparison']
        
        # Header formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Highlight average row
        avg_row = len(glitch_df)
        avg_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        for cell in ws[avg_row + 1]:
            cell.fill = avg_fill
            cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width


def main():
    """Main function for command-line usage"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate Excel comparison report")
    parser.add_argument("--iac-tech", default="puppet", help="IaC technology")
    parser.add_argument("--experiment-ids", nargs="+", help="Specific experiment IDs")
    parser.add_argument("--latest", type=int, default=2, help="Use latest N experiments")
    parser.add_argument("--output", help="Output Excel file path")
    
    args = parser.parse_args()
    
    generator = ExperimentReportGenerator()
    
    # Determine experiment IDs to use
    if args.experiment_ids:
        experiment_ids = args.experiment_ids
    else:
        experiment_ids = generator.find_latest_experiments(args.iac_tech, args.latest)
        print(f"Using latest {args.latest} experiments: {experiment_ids}")
    
    # Generate report
    try:
        output_file = generator.generate_excel_report(
            experiment_ids, 
            args.iac_tech, 
            args.output
        )
        print(f"✅ Report generation complete: {output_file}")
    except Exception as e:
        print(f"❌ Failed to generate report: {e}")
        return 1
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
